#Archivo de reportes en Excel

# Archivo de reportes en Excel mejorado
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_excel(data, filename="reporte.xlsx"):
    """
    Genera un reporte en Excel con formato mejorado a partir de datos JSON/SQL/NoSQL.
    """
    # Convertir en DataFrame
    df = pd.DataFrame(data)

    # Guardar en Excel (con soporte de openpyxl para aplicar estilos)
    df.to_excel(filename, index=False, engine="openpyxl")

    # Abrir libro para aplicar estilos
    wb = load_workbook(filename)
    ws = wb.active

    # Encabezados con estilo
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")

    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

        # Ajuste automático del ancho de columna
        max_length = max(
            (len(str(value)) for value in df[column_title].astype(str)),
            default=len(column_title)
        )
        adjusted_width = max(max_length + 2, len(column_title) + 2)
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # Ajuste general de alineación para todo el contenido
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Guardar cambios
    wb.save(filename)
    return filename